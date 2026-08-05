import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MONEY = "#,##0.00"
WHOLE = "#,##0"
PERCENT = "0.00"

DEAL_COLUMNS = [
    ("Company Name", "company_name", 34, None),
    ("Trade Date", "trade_date_iso", 13, "DD-MMM-YYYY"),
    ("Year", "year", 7, None),
    ("Ticker", "ticker", 13, None),
    ("Quantity Traded", "quantity", 16, WHOLE),
    ("Trade Price / Wtd. Avg. Price", "price", 22, None),
    ("Deal Size (Rs cr)", "deal_size_cr", 15, MONEY),
    ("Prev Close (Rs)", "prev_close", 14, MONEY),
    ("Discount to Prev Close (%)", "discount_pct", 16, PERCENT),
    # The month of trading up to the session before the deal.
    ("1M Return before Deal (%)", "pre_return_1m_pct", 16, PERCENT),
    ("1M ADTV before Deal (Rs cr)", "pre_adtv_1m_cr", 17, MONEY),
    ("1M VWAP before Deal (Rs)", "pre_vwap_1m", 16, MONEY),
    ("1M Delivery before Deal (%)", "pre_delivery_1m_pct", 17, PERCENT),
    ("Type", "deal_type", 20, None),
    ("Sellers", "sellers", 60, None),
    ("Buyers", "buyers", 60, None),
    ("Exchange", "exchange", 12, None),
]

MARKET_BASE_COLUMNS = [
    ("Company Name", "company_name", 34, None),
    ("CMP (Rs)", "cmp", 12, MONEY),
    ("Intraday (%)", "intraday_return_pct", 12, PERCENT),
    ("Daily (%)", "daily_return_pct", 12, PERCENT),
    ("Market Cap (Rs cr)", "market_cap_cr", 17, MONEY),
    ("Return 1D (%)", "return_1d_pct", 13, PERCENT),
    ("Return 1W (%)", "return_1w_pct", 13, PERCENT),
    ("Return 1M (%)", "return_1m_pct", 13, PERCENT),
]

# ADTV / VWAP / Delivery follow the Market-data window checkboxes; returns do not.
MARKET_VOLUME_COLUMNS = {
    "1d": [
        ("ADTV 1D (Rs cr)", "adtv_1d_cr", 15, MONEY),
        ("VWAP 1D (Rs)", "vwap_1d", 13, MONEY),
        ("Delivery 1D (%)", "delivery_1d_pct", 15, PERCENT),
    ],
    "1w": [
        ("ADTV 1W (Rs cr)", "adtv_1w_cr", 15, MONEY),
        ("VWAP 1W (Rs)", "vwap_1w", 13, MONEY),
        ("Delivery 1W (%)", "delivery_1w_pct", 15, PERCENT),
    ],
    "1m": [
        ("ADTV 1M (Rs cr)", "adtv_1m_cr", 15, MONEY),
        ("VWAP 1M (Rs)", "vwap_1m", 13, MONEY),
        ("Delivery 1M (%)", "delivery_1m_pct", 15, PERCENT),
    ],
}


def market_columns(windows=None) -> list:
    """Return % always; ADTV/VWAP/Delivery only for the chosen windows."""
    wanted = windows or {"1d": True, "1w": True, "1m": True}
    columns = list(MARKET_BASE_COLUMNS)
    # Keep metric families together (all ADTVs, then VWAPs, then delivery).
    for metric_index in range(3):
        for period in ("1d", "1w", "1m"):
            if wanted.get(period, True):
                columns.append(MARKET_VOLUME_COLUMNS[period][metric_index])
    return columns


# Default full set — used by tests / callers that do not pass windows.
MARKET_COLUMNS = market_columns()

NEWS_COLUMNS = [
    ("Company Name", "company_name", 34, None),
    ("Source", "source", 18, None),
    ("Published", "published_display", 17, None),
    ("Headline", "headline", 90, None),
    ("Link", "url", 60, None),
]

QUARTER_COLUMNS = [
    ("Company Name", "company_name", 34, None),
    ("Ticker", "ticker", 13, None),
    ("Quarter", "quarter", 22, None),
    ("Period Ended", "period_end", 14, None),
    ("Basis", "consolidated", 14, None),
    ("Key takeaways", "takeaways_text", 80, None),
    ("Report", "report_label", 16, None),
    ("Report link", "report_url", 60, None),
]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
LINK_FONT = Font(color="1155CC", underline="single")
# Intraday / Daily beyond ±3% — whole market row (matches the PDF).
ROW_HIGHLIGHT_FILL = PatternFill("solid", fgColor="DBEAFE")
COMPANY_BOLD = Font(bold=True)
LIVE_MOVE_THRESHOLD = 3.0


def _as_date(value):
    try:
        return date.fromisoformat(value)
    except (TypeError, ValueError):
        return value


def _big_live_move(record) -> bool:
    """True when Intraday % or Daily % is outside ±3%."""
    for field in ("intraday_return_pct", "daily_return_pct"):
        value = record.get(field)
        if value is None:
            continue
        try:
            if abs(float(value)) > LIVE_MOVE_THRESHOLD:
                return True
        except (TypeError, ValueError):
            continue
    return False


def _write_sheet(sheet, columns, records, highlight_big_moves=False) -> None:
    sheet.append([title for title, _, _, _ in columns])
    for index, (_, _, width, _) in enumerate(columns, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = width
        cell = sheet.cell(row=1, column=index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30

    for record in records:
        values = []
        for _, field, _, number_format in columns:
            value = record.get(field)
            values.append(_as_date(value) if number_format == "DD-MMM-YYYY" else value)
        sheet.append(values)
        row_index = sheet.max_row
        big = highlight_big_moves and _big_live_move(record)
        for index, (_, field, _, number_format) in enumerate(columns, start=1):
            cell = sheet.cell(row=row_index, column=index)
            if number_format:
                cell.number_format = number_format
            if big:
                cell.fill = ROW_HIGHLIGHT_FILL
                if field == "company_name":
                    cell.font = COMPANY_BOLD

    last_row = sheet.max_row
    sheet.freeze_panes = "A2"
    if last_row > 1:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{last_row}"


def _news_records(articles):
    records = []
    for article in articles:
        record = dict(article)
        published = article.get("published") or ""
        record["published_display"] = published.replace("T", " ")[:16]
        records.append(record)
    return records


def _write_news(sheet, articles) -> None:
    _write_sheet(sheet, NEWS_COLUMNS, _news_records(articles))
    headline_column = 1 + [field for _, field, _, _ in NEWS_COLUMNS].index("headline")
    link_column = 1 + [field for _, field, _, _ in NEWS_COLUMNS].index("url")

    for row_index, article in enumerate(articles, start=2):
        url = article.get("url")
        if not url:
            continue
        for column in (headline_column, link_column):
            cell = sheet.cell(row=row_index, column=column)
            cell.hyperlink = url
            cell.font = LINK_FONT


def _quarter_records(entries):
    records = []
    for entry in entries or []:
        record = dict(entry)
        record["takeaways_text"] = "\n".join(f"• {point}" for point in entry.get("takeaways") or [])
        records.append(record)
    return records


def _write_quarters(sheet, entries) -> None:
    records = _quarter_records(entries)
    _write_sheet(sheet, QUARTER_COLUMNS, records)
    takeaways_column = 1 + [field for _, field, _, _ in QUARTER_COLUMNS].index("takeaways_text")
    link_column = 1 + [field for _, field, _, _ in QUARTER_COLUMNS].index("report_url")
    label_column = 1 + [field for _, field, _, _ in QUARTER_COLUMNS].index("report_label")

    for row_index, entry in enumerate(entries or [], start=2):
        sheet.cell(row=row_index, column=takeaways_column).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        points = entry.get("takeaways") or []
        if points:
            sheet.row_dimensions[row_index].height = max(45, 14 * len(points) + 10)
        url = entry.get("report_url")
        if not url:
            continue
        for column in (label_column, link_column):
            cell = sheet.cell(row=row_index, column=column)
            cell.hyperlink = url
            cell.font = LINK_FONT


def build_workbook(
    rows,
    market_data=None,
    articles=None,
    quarters=None,
    sections=None,
    market_windows=None,
) -> bytes:
    """Build only the sheets the run asked for."""
    wanted = sections or {
        "deals": True,
        "market": True,
        "news": True,
        "quarters": True,
    }
    workbook = Workbook()
    first = workbook.active
    used_first = False

    def sheet(title: str):
        nonlocal used_first
        if not used_first:
            first.title = title
            used_first = True
            return first
        return workbook.create_sheet(title)

    if wanted.get("deals", True):
        deals = sheet("Block & Bulk Deals")
        _write_sheet(deals, DEAL_COLUMNS, rows or [])
        fields = [field for _, field, _, _ in DEAL_COLUMNS]
        parties = [1 + fields.index(name) for name in ("sellers", "buyers")]
        for row_index in range(2, deals.max_row + 1):
            for column_index in parties:
                deals.cell(row=row_index, column=column_index).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

    if wanted.get("market", True) and market_data:
        _write_sheet(
            sheet("Market Data"),
            market_columns(market_windows),
            market_data,
            highlight_big_moves=True,
        )
    if wanted.get("quarters", True) and quarters:
        _write_quarters(sheet("Quarterly Results"), quarters)

    if wanted.get("news", True) and articles:
        _write_news(sheet("News"), articles)

    if not used_first:
        first.title = "Tracker"
        first["A1"] = "No sections were selected for this run."

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()
